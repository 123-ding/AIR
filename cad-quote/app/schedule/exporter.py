"""配电箱回路表导出工具。"""

from __future__ import annotations

import csv
import json
import os
from dataclasses import asdict
from typing import List

from .models import PanelSchedule


def schedule_to_dict(schedule: PanelSchedule) -> dict:
    return asdict(schedule)


def schedule_to_rows(schedule: PanelSchedule) -> List[dict]:
    rows = []
    for circuit in schedule.circuits:
        rows.append(
            {
                "回路": circuit.circuit,
                "断路器": circuit.breaker,
                "极数": circuit.poles,
                "曲线": circuit.curve,
                "整定": circuit.rating,
                "相序": circuit.phase,
                "电缆": circuit.cable,
                "敷设": circuit.conduit,
                "负荷": circuit.load,
                "用途": circuit.usage,
            }
        )
    return rows


def schedule_to_dataframe(schedule: PanelSchedule):
    try:
        import pandas as pd  # type: ignore
    except ImportError as exc:  # pragma: no cover
        raise ImportError("pandas 未安装；请 `pip install pandas`。") from exc
    return pd.DataFrame(schedule_to_rows(schedule))


def to_json(schedule: PanelSchedule, out_path: str) -> str:
    os.makedirs(os.path.dirname(os.path.abspath(out_path)) or ".", exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(schedule_to_dict(schedule), f, ensure_ascii=False, indent=2)
    return out_path


def to_csv(schedule: PanelSchedule, out_path: str) -> str:
    os.makedirs(os.path.dirname(os.path.abspath(out_path)) or ".", exist_ok=True)
    header_rows = [
        ("箱名", schedule.header.name),
        ("编号", schedule.header.code),
        ("Pe", schedule.header.pe),
        ("Kx", schedule.header.kx),
        ("cosφ", schedule.header.cos_phi),
        ("Ijs", schedule.header.ijs),
        ("进线总开关", schedule.header.main_breaker),
        ("接触器", schedule.header.contactor),
        ("SPD", schedule.header.spd),
        ("尺寸", schedule.header.size),
        ("安装方式", schedule.header.install),
    ]
    rows = schedule_to_rows(schedule)
    fieldnames = ["回路", "断路器", "极数", "曲线", "整定", "相序", "电缆", "敷设", "负荷", "用途"]
    with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["箱头字段", "值"])
        for key, value in header_rows:
            writer.writerow([key, value])
        for key, value in schedule.header.extras.items():
            writer.writerow([key, value])
        writer.writerow([])
        writer.writerow(fieldnames)
        dict_writer = csv.DictWriter(f, fieldnames=fieldnames)
        dict_writer.writerows(rows)
    return out_path


def to_excel(
    schedule: PanelSchedule,
    out_path: str,
    circuit_sheet_name: str = "回路表",
    header_sheet_name: str = "配电箱头",
) -> str:
    try:
        from openpyxl import Workbook  # type: ignore
        from openpyxl.styles import Alignment, Font  # type: ignore
    except ImportError as exc:  # pragma: no cover
        raise ImportError("openpyxl 未安装；请 `pip install openpyxl`。") from exc

    os.makedirs(os.path.dirname(os.path.abspath(out_path)) or ".", exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = circuit_sheet_name

    header = ["回路", "断路器", "极数", "曲线", "整定", "相序", "电缆", "敷设", "负荷", "用途"]
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    for row in schedule_to_rows(schedule):
        ws.append([row[key] for key in header])
    widths = [10, 18, 8, 8, 14, 8, 20, 10, 10, 28]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width

    header_ws = wb.create_sheet(header_sheet_name)
    header_ws.append(["字段", "值"])
    for cell in header_ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    header_rows = [
        ("箱名", schedule.header.name),
        ("编号", schedule.header.code),
        ("Pe", schedule.header.pe),
        ("Kx", schedule.header.kx),
        ("cosφ", schedule.header.cos_phi),
        ("Ijs", schedule.header.ijs),
        ("进线总开关", schedule.header.main_breaker),
        ("接触器", schedule.header.contactor),
        ("SPD", schedule.header.spd),
        ("尺寸", schedule.header.size),
        ("安装方式", schedule.header.install),
    ]
    for row in header_rows:
        header_ws.append(list(row))
    for key, value in schedule.header.extras.items():
        header_ws.append([key, value])
    header_ws.column_dimensions["A"].width = 18
    header_ws.column_dimensions["B"].width = 48

    wb.save(out_path)
    return out_path
